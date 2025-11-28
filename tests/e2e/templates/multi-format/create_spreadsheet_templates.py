#!/usr/bin/env python3
"""
Create Excel and LibreOffice Calc template files

This script generates .xlsx and .ods template files for the AI Assistant
to accept and parse when creating organizations.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from odf.opendocument import OpenDocumentSpreadsheet
from odf.style import Style, TextProperties, ParagraphProperties, TableCellProperties
from odf.text import P
from odf.table import Table, TableRow, TableCell


def create_excel_template():
    """Create Excel (.xlsx) template file"""
    print("Creating Excel template...")

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Organization
    ws_org = wb.create_sheet("Organization")
    headers_org = [
        "name", "slug", "description", "domain", "industry", "size",
        "street", "city", "state", "postal_code", "country",
        "phone", "email", "website",
        "admin_full_name", "admin_username", "admin_email", "admin_password",
        "admin_first_name", "admin_last_name", "admin_role",
        "timezone", "language", "currency", "date_format"
    ]

    for col, header in enumerate(headers_org, 1):
        cell = ws_org.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add example data
    data_org = [
        "TechEd Academy", "teched-academy",
        "Modern technology education academy focusing on software development",
        "teched-academy.edu", "Education & Technology", "small",
        "456 Innovation Drive", "San Francisco", "CA", "94105", "US",
        "+1-415-555-0199", "contact@teched-academy.edu", "https://teched-academy.edu",
        "Dr. Emily Chen", "echen", "admin@teched-academy.edu", "AdminPass123!",
        "Emily", "Chen", "organization_admin",
        "America/Los_Angeles", "en", "USD", "MM/DD/YYYY"
    ]

    for col, value in enumerate(data_org, 1):
        ws_org.cell(row=2, column=col, value=value)

    # Auto-size columns
    for col in ws_org.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_org.column_dimensions[column].width = adjusted_width

    # Sheet 2: Instructors
    ws_inst = wb.create_sheet("Instructors")
    headers_inst = [
        "first_name", "last_name", "email", "full_name", "username",
        "password", "specialties", "bio"
    ]

    for col, header in enumerate(headers_inst, 1):
        cell = ws_inst.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add instructor data
    instructors = [
        ["Alex", "Rodriguez", "alex.rodriguez@teched-academy.edu", "Alex Rodriguez",
         "arodriguez", "InstructorPass123!", "React, JavaScript",
         "Senior frontend developer with 8 years of experience"],
        ["Maya", "Patel", "maya.patel@teched-academy.edu", "Maya Patel",
         "mpatel", "InstructorPass123!", "Node.js, Databases",
         "Backend engineer specializing in scalable systems"]
    ]

    for row_idx, instructor in enumerate(instructors, 2):
        for col_idx, value in enumerate(instructor, 1):
            ws_inst.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col in ws_inst.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_inst.column_dimensions[column].width = adjusted_width

    # Sheet 3: Students
    ws_stud = wb.create_sheet("Students")
    headers_stud = [
        "first_name", "last_name", "email", "full_name", "username",
        "password", "background", "goals"
    ]

    for col, header in enumerate(headers_stud, 1):
        cell = ws_stud.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add student data
    students = [
        ["Jordan", "Smith", "jordan.smith@students.teched-academy.edu", "Jordan Smith",
         "jsmith", "StudentPass123!", "Career changer from marketing",
         "Become a frontend developer, Build portfolio projects"],
        ["Taylor", "Johnson", "taylor.johnson@students.teched-academy.edu", "Taylor Johnson",
         "tjohnson", "StudentPass123!", "Computer science graduate",
         "Master full-stack development, Get hired at a tech company"],
        ["Casey", "Williams", "casey.williams@students.teched-academy.edu", "Casey Williams",
         "cwilliams", "StudentPass123!", "Self-taught developer",
         "Formalize web development skills, Learn best practices"]
    ]

    for row_idx, student in enumerate(students, 2):
        for col_idx, value in enumerate(student, 1):
            ws_stud.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col in ws_stud.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_stud.column_dimensions[column].width = adjusted_width

    # Sheet 4: Tracks
    ws_tracks = wb.create_sheet("Tracks")
    headers_tracks = [
        "name", "description", "target_audience", "prerequisites",
        "learning_objectives", "duration_weeks", "difficulty", "max_students"
    ]

    for col, header in enumerate(headers_tracks, 1):
        cell = ws_tracks.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add track data
    tracks = [
        ["Frontend Development", "Master modern frontend technologies",
         "career_changers, junior_developers",
         "Basic HTML/CSS knowledge, JavaScript fundamentals",
         "Build responsive web applications with React, Implement state management with Redux, Create accessible and performant UIs",
         6, "beginner", 30],
        ["Backend Development", "Server-side development with Node.js",
         "frontend_developers, junior_developers",
         "JavaScript proficiency, Understanding of HTTP",
         "Build RESTful APIs with Express, Work with databases (SQL and NoSQL), Implement authentication and authorization",
         6, "intermediate", 30]
    ]

    for row_idx, track in enumerate(tracks, 2):
        for col_idx, value in enumerate(track, 1):
            ws_tracks.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col in ws_tracks.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_tracks.column_dimensions[column].width = adjusted_width

    # Sheet 5: Courses
    ws_courses = wb.create_sheet("Courses")
    headers_courses = ["track", "title", "description", "difficulty", "duration_weeks"]

    for col, header in enumerate(headers_courses, 1):
        cell = ws_courses.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add course data
    courses = [
        ["Frontend Development", "React Fundamentals",
         "Learn React basics including components, props, and state", "beginner", 2],
        ["Frontend Development", "Advanced React Patterns",
         "Master hooks, context, and performance optimization", "intermediate", 2],
        ["Frontend Development", "State Management with Redux",
         "Implement complex state management in React apps", "intermediate", 2],
        ["Backend Development", "Node.js Essentials",
         "Build server-side applications with Node.js", "beginner", 2],
        ["Backend Development", "Express.js and REST APIs",
         "Create RESTful APIs with Express framework", "intermediate", 2],
        ["Backend Development", "Database Integration",
         "Work with PostgreSQL and MongoDB", "intermediate", 2]
    ]

    for row_idx, course in enumerate(courses, 2):
        for col_idx, value in enumerate(course, 1):
            ws_courses.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col in ws_courses.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_courses.column_dimensions[column].width = adjusted_width

    # Save workbook
    output_path = "organization_simple.xlsx"
    wb.save(output_path)
    print(f"✅ Excel template created: {output_path}")


def create_ods_template():
    """Create LibreOffice Calc (.ods) template file"""
    print("Creating LibreOffice Calc template...")

    # Create spreadsheet
    doc = OpenDocumentSpreadsheet()

    # Create organization table
    org_table = Table(name="Organization")

    # Add header row
    headers_org = [
        "name", "slug", "description", "domain", "industry", "size",
        "street", "city", "state", "postal_code", "country",
        "phone", "email", "website",
        "admin_full_name", "admin_username", "admin_email", "admin_password",
        "admin_first_name", "admin_last_name", "admin_role",
        "timezone", "language", "currency", "date_format"
    ]

    header_row = TableRow()
    for header in headers_org:
        cell = TableCell()
        cell.addElement(P(text=header))
        header_row.addElement(cell)
    org_table.addElement(header_row)

    # Add data row
    data_org = [
        "TechEd Academy", "teched-academy",
        "Modern technology education academy focusing on software development",
        "teched-academy.edu", "Education & Technology", "small",
        "456 Innovation Drive", "San Francisco", "CA", "94105", "US",
        "+1-415-555-0199", "contact@teched-academy.edu", "https://teched-academy.edu",
        "Dr. Emily Chen", "echen", "admin@teched-academy.edu", "AdminPass123!",
        "Emily", "Chen", "organization_admin",
        "America/Los_Angeles", "en", "USD", "MM/DD/YYYY"
    ]

    data_row = TableRow()
    for value in data_org:
        cell = TableCell()
        cell.addElement(P(text=str(value)))
        data_row.addElement(cell)
    org_table.addElement(data_row)

    doc.spreadsheet.addElement(org_table)

    # Save document
    output_path = "organization_simple.ods"
    doc.save(output_path)
    print(f"✅ LibreOffice Calc template created: {output_path}")


if __name__ == "__main__":
    try:
        create_excel_template()
        create_ods_template()
        print("\n✅ All spreadsheet templates created successfully!")
    except ImportError as e:
        print(f"⚠️  Missing dependency: {e}")
        print("Install required packages: pip install openpyxl odfpy")
    except Exception as e:
        print(f"❌ Error creating templates: {e}")
        import traceback
        traceback.print_exc()
