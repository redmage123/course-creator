"""
Unit Tests for Project Spreadsheet Parser

BUSINESS CONTEXT:
Tests the project spreadsheet parser service that enables organization admins
to upload spreadsheets for automated project creation with tracks, students,
and instructors.

TEST COVERAGE:
- Excel (XLSX/XLS) file parsing
- LibreOffice Calc (ODS) file parsing
- Required field validation
- Optional field handling (dates, tracks, students, instructors)
- Email parsing and validation
- Name parsing for students/instructors
- Template generation
- Error handling and validation messages
"""
import pytest
import io
import openpyxl
from openpyxl import Workbook
from datetime import datetime

# Import the parser
import sys
sys.path.insert(0, '/home/bbrelin/course-creator/services/course-management')
from course_management.application.services.project_spreadsheet_parser import ProjectSpreadsheetParser


class TestProjectSpreadsheetParser:
    """Test suite for ProjectSpreadsheetParser service"""

    def setup_method(self):
        """Set up test fixtures"""
        self.parser = ProjectSpreadsheetParser()

    def create_test_xlsx(self, data: dict) -> bytes:
        """Helper to create test XLSX file"""
        wb = Workbook()
        ws = wb.active

        # Write headers
        headers = list(data.keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data row
        values = list(data.values())
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    # ========================================================================
    # REQUIRED FIELDS TESTS
    # ========================================================================

    def test_parse_minimal_project_data(self):
        """Test parsing spreadsheet with only required fields"""
        data = {
            'project_name': 'Python Training 2024',
            'project_slug': 'python-training-2024',
            'description': 'Comprehensive Python training program'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert result['project_name'] == 'Python Training 2024'
        assert result['project_slug'] == 'python-training-2024'
        assert result['description'] == 'Comprehensive Python training program'
        assert result['students'] == []
        assert result['instructors'] == []
        assert result['tracks'] == []

    def test_missing_required_field_raises_error(self):
        """Test that missing required fields raise ValueError"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training'
            # Missing description
        }

        xlsx_bytes = self.create_test_xlsx(data)

        with pytest.raises(ValueError) as exc_info:
            self.parser.parse_xlsx(xlsx_bytes)

        assert 'Missing required columns' in str(exc_info.value)
        assert 'description' in str(exc_info.value)

    def test_empty_required_field_raises_error(self):
        """Test that empty required fields raise ValueError"""
        data = {
            'project_name': '',  # Empty
            'project_slug': 'python-training',
            'description': 'Test description'
        }

        xlsx_bytes = self.create_test_xlsx(data)

        with pytest.raises(ValueError) as exc_info:
            self.parser.parse_xlsx(xlsx_bytes)

        assert 'project_name cannot be empty' in str(exc_info.value)

    # ========================================================================
    # OPTIONAL FIELDS TESTS
    # ========================================================================

    def test_parse_with_dates(self):
        """Test parsing project with start and end dates"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'start_date': '2024-01-15',
            'end_date': '2024-06-30'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert result['start_date'] == '2024-01-15'
        assert result['end_date'] == '2024-06-30'

    def test_parse_with_tracks(self):
        """Test parsing project with comma-separated tracks"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'tracks': 'Backend Fundamentals, Frontend Basics, Database Design'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert len(result['tracks']) == 3
        assert 'Backend Fundamentals' in result['tracks']
        assert 'Frontend Basics' in result['tracks']
        assert 'Database Design' in result['tracks']

    # ========================================================================
    # STUDENTS PARSING TESTS
    # ========================================================================

    def test_parse_students_with_names(self):
        """Test parsing students with emails and names"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'student_emails': 'alice@example.com, bob@example.com',
            'student_names': 'Alice Johnson, Bob Smith'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert len(result['students']) == 2
        assert result['students'][0] == {'email': 'alice@example.com', 'name': 'Alice Johnson'}
        assert result['students'][1] == {'email': 'bob@example.com', 'name': 'Bob Smith'}

    def test_parse_students_without_names(self):
        """Test parsing students with only emails"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'student_emails': 'alice@example.com, bob@example.com'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert len(result['students']) == 2
        assert result['students'][0] == {'email': 'alice@example.com'}
        assert result['students'][1] == {'email': 'bob@example.com'}

    def test_parse_students_mismatched_names(self):
        """Test handling when email count doesn't match name count"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'student_emails': 'alice@example.com, bob@example.com, charlie@example.com',
            'student_names': 'Alice Johnson, Bob Smith'  # Only 2 names for 3 emails
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert len(result['students']) == 3
        assert result['students'][0] == {'email': 'alice@example.com', 'name': 'Alice Johnson'}
        assert result['students'][1] == {'email': 'bob@example.com', 'name': 'Bob Smith'}
        assert result['students'][2] == {'email': 'charlie@example.com'}  # No name

    # ========================================================================
    # INSTRUCTORS PARSING TESTS
    # ========================================================================

    def test_parse_instructors_with_names(self):
        """Test parsing instructors with emails and names"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'instructor_emails': 'dr.smith@example.com',
            'instructor_names': 'Dr. Sarah Smith'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert len(result['instructors']) == 1
        assert result['instructors'][0] == {'email': 'dr.smith@example.com', 'name': 'Dr. Sarah Smith'}

    def test_parse_multiple_instructors(self):
        """Test parsing multiple instructors"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'instructor_emails': 'instructor1@example.com, instructor2@example.com',
            'instructor_names': 'John Doe, Jane Smith'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert len(result['instructors']) == 2
        assert result['instructors'][0]['email'] == 'instructor1@example.com'
        assert result['instructors'][1]['email'] == 'instructor2@example.com'

    # ========================================================================
    # COMPLETE PROJECT TESTS
    # ========================================================================

    def test_parse_complete_project(self):
        """Test parsing complete project with all fields"""
        data = {
            'project_name': 'Python Web Development Training',
            'project_slug': 'python-web-dev-2024',
            'description': 'Comprehensive training program',
            'start_date': '2024-01-15',
            'end_date': '2024-06-30',
            'tracks': 'Backend, Frontend, Database',
            'student_emails': 'student1@example.com, student2@example.com',
            'student_names': 'Alice Johnson, Bob Smith',
            'instructor_emails': 'instructor@example.com',
            'instructor_names': 'Dr. Sarah Chen'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        # Verify all fields
        assert result['project_name'] == 'Python Web Development Training'
        assert result['project_slug'] == 'python-web-dev-2024'
        assert len(result['tracks']) == 3
        assert len(result['students']) == 2
        assert len(result['instructors']) == 1
        assert result['start_date'] == '2024-01-15'
        assert result['end_date'] == '2024-06-30'

    # ========================================================================
    # TEMPLATE GENERATION TESTS
    # ========================================================================

    def test_generate_template(self):
        """Test template generation produces valid XLSX"""
        template_bytes = ProjectSpreadsheetParser.generate_template()

        assert template_bytes is not None
        assert len(template_bytes) > 0

        # Verify it's valid XLSX by parsing it
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb.active

        # Check headers are present
        headers = [cell.value for cell in ws[1]]
        assert 'project_name' in headers
        assert 'project_slug' in headers
        assert 'description' in headers
        assert 'student_emails' in headers
        assert 'instructor_emails' in headers

        # Check example data is present
        assert ws['A2'].value == 'Python Web Development Training'

    def test_generated_template_is_parseable(self):
        """Test that generated template can be parsed successfully"""
        template_bytes = ProjectSpreadsheetParser.generate_template()

        # Parse the template
        result = self.parser.parse_xlsx(template_bytes)

        # Should have example data
        assert result['project_name'] == 'Python Web Development Training'
        assert len(result['tracks']) == 3
        assert len(result['students']) == 2
        assert len(result['instructors']) == 1

    # ========================================================================
    # FORMAT DETECTION TESTS
    # ========================================================================

    def test_detect_xlsx_format(self):
        """Test format detection for XLSX files"""
        assert self.parser.detect_format('project.xlsx') == 'xlsx'
        assert self.parser.detect_format('project.xls') == 'xlsx'
        assert self.parser.detect_format('PROJECT.XLSX') == 'xlsx'

    def test_detect_ods_format(self):
        """Test format detection for ODS files"""
        assert self.parser.detect_format('project.ods') == 'ods'
        assert self.parser.detect_format('PROJECT.ODS') == 'ods'

    def test_unsupported_format_raises_error(self):
        """Test that unsupported formats raise ValueError"""
        with pytest.raises(ValueError) as exc_info:
            self.parser.detect_format('project.pdf')

        assert 'Unsupported file type' in str(exc_info.value)

    # ========================================================================
    # ERROR HANDLING TESTS
    # ========================================================================

    def test_empty_spreadsheet_raises_error(self):
        """Test that empty spreadsheet raises ValueError"""
        with pytest.raises(ValueError) as exc_info:
            self.parser.parse_xlsx(b'')

        assert 'Empty spreadsheet' in str(exc_info.value)

    def test_invalid_xlsx_raises_error(self):
        """Test that invalid XLSX data raises ValueError"""
        invalid_data = b'This is not valid XLSX data'

        with pytest.raises(ValueError) as exc_info:
            self.parser.parse_xlsx(invalid_data)

        assert 'Failed to parse XLSX' in str(exc_info.value)

    # ========================================================================
    # DATA CLEANING TESTS
    # ========================================================================

    def test_whitespace_trimming(self):
        """Test that whitespace is properly trimmed"""
        data = {
            'project_name': '  Python Training  ',
            'project_slug': '  python-training  ',
            'description': '  Test description  ',
            'tracks': ' Backend , Frontend , Database '
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert result['project_name'] == 'Python Training'
        assert result['project_slug'] == 'python-training'
        assert result['description'] == 'Test description'
        assert 'Backend' in result['tracks']
        assert 'Frontend' in result['tracks']

    def test_slug_lowercase_conversion(self):
        """Test that project slug is converted to lowercase"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'PYTHON-TRAINING-2024',
            'description': 'Test description'
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert result['project_slug'] == 'python-training-2024'

    def test_empty_tracks_list_handled(self):
        """Test that empty tracks string results in empty list"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'tracks': '   '  # Only whitespace
        }

        xlsx_bytes = self.create_test_xlsx(data)
        result = self.parser.parse_xlsx(xlsx_bytes)

        assert result['tracks'] == []


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
