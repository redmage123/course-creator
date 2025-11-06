"""
Integration Tests for Project Import API Endpoints

BUSINESS CONTEXT:
Tests the project import API endpoints that enable organization admins to:
1. Upload spreadsheets for parsing
2. Download template spreadsheets
3. Create projects automatically from validated data

TEST COVERAGE:
- POST /api/v1/projects/import-spreadsheet endpoint
- GET /api/v1/projects/template endpoint
- POST /api/v1/projects/create-from-spreadsheet endpoint
- Authentication and authorization
- File upload validation
- Error handling and edge cases
"""
import pytest
import asyncio
import io
import openpyxl
from openpyxl import Workbook
from fastapi.testclient import TestClient

# Note: Path setup is handled by conftest.py
# Infrastructure fixtures provide test_app and auth_headers


class TestProjectImportAPI:
    """
    Integration tests for project import API endpoints with real infrastructure

    BUSINESS CONTEXT:
    Tests the complete API workflow with actual database and Redis connections.
    Uses docker-compose.test.yml infrastructure (PostgreSQL + Redis).

    FIXTURES USED:
    - test_app: FastAPI TestClient with real infrastructure
    - auth_headers: Valid JWT authentication headers
    - clean_database: Ensures clean database state per test
    - test_project_data: Sample project data for testing
    """

    def create_test_xlsx(self, data: dict) -> bytes:
        """Helper to create test XLSX file"""
        wb = Workbook()
        ws = wb.active

        # Write headers
        headers = list(data.keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data row
        values = list(data.values())
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    # ========================================================================
    # IMPORT SPREADSHEET ENDPOINT TESTS
    # ========================================================================

    def test_import_spreadsheet_success(self, test_app, auth_headers, clean_database):
        """Test successful spreadsheet import with real infrastructure"""
        data = {
            'project_name': 'Python Training 2024',
            'project_slug': 'python-training-2024',
            'description': 'Comprehensive Python training',
            'tracks': 'Backend, Frontend',
            'student_emails': 'student1@example.com, student2@example.com',
            'student_names': 'Alice Johnson, Bob Smith'
        }

        xlsx_bytes = self.create_test_xlsx(data)

        response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            headers=auth_headers,
            files={"file": ("project.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        result = response.json()
        assert result['success'] is True
        assert result['project_name'] == 'Python Training 2024'
        assert result['project_slug'] == 'python-training-2024'
        assert len(result['tracks']) == 2
        assert len(result['students']) == 2

    def test_import_spreadsheet_minimal_data(self, test_app, auth_headers, clean_database):
        """Test import with only required fields"""
        data = {
            'project_name': 'Minimal Project',
            'project_slug': 'minimal-project',
            'description': 'Test description'
        }

        xlsx_bytes = self.create_test_xlsx(data)

        response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            headers=auth_headers,
            files={"file": ("project.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 200
        result = response.json()
        assert result['success'] is True
        assert result['students'] == []
        assert result['instructors'] == []
        assert result['tracks'] == []

    def test_import_spreadsheet_missing_required_field(self, test_app, auth_headers, clean_database):
        """Test import with missing required field returns 400"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training'
            # Missing description
        }

        xlsx_bytes = self.create_test_xlsx(data)

        response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            headers=auth_headers,
            files={"file": ("project.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 400
        assert 'Missing required columns' in response.json()['detail']

    def test_import_spreadsheet_unsupported_format(self, test_app, auth_headers, clean_database):
        """Test upload of unsupported file format returns 400"""
        response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            headers=auth_headers,
            files={"file": ("project.pdf", b"fake pdf content", "application/pdf")}
        )

        assert response.status_code == 400
        assert 'Unsupported file type' in response.json()['detail']

    def test_import_spreadsheet_file_too_large(self, test_app, auth_headers, clean_database):
        """Test upload of oversized file returns 413"""
        # Create a file larger than 10MB
        large_data = b"x" * (11 * 1024 * 1024)  # 11MB

        response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            headers=auth_headers,
            files={"file": ("large.xlsx", large_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code == 413
        assert 'File too large' in response.json()['detail']

    def test_import_spreadsheet_without_auth(self, test_app, auth_headers, clean_database):
        """Test upload without authentication returns 401 or 403"""
        data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description'
        }

        xlsx_bytes = self.create_test_xlsx(data)

        response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            files={"file": ("project.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        # Should require authentication
        assert response.status_code in [401, 403, 422]  # Depends on auth implementation

    # ========================================================================
    # TEMPLATE DOWNLOAD ENDPOINT TESTS
    # ========================================================================

    def test_download_template_success(self, test_app, auth_headers, clean_database):
        """Test successful template download"""
        response = test_app.get(
            "/api/v1/projects/template",
            headers=auth_headers
        )

        assert response.status_code == 200
        assert response.headers['content-type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'attachment' in response.headers['content-disposition']
        assert 'project_import_template.xlsx' in response.headers['content-disposition']

        # Verify content is valid XLSX
        assert len(response.content) > 0
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        assert wb.active is not None

    def test_download_template_has_correct_headers(self, test_app, auth_headers, clean_database):
        """Test template has all required and optional headers"""
        response = test_app.get(
            "/api/v1/projects/template",
            headers=auth_headers
        )

        assert response.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Check all required headers
        assert 'project_name' in headers
        assert 'project_slug' in headers
        assert 'description' in headers

        # Check optional headers
        assert 'start_date' in headers
        assert 'end_date' in headers
        assert 'tracks' in headers
        assert 'student_emails' in headers
        assert 'student_names' in headers
        assert 'instructor_emails' in headers
        assert 'instructor_names' in headers

    def test_download_template_has_example_data(self, test_app, auth_headers, clean_database):
        """Test template includes example data in row 2"""
        response = test_app.get(
            "/api/v1/projects/template",
            headers=auth_headers
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Check example data is present
        assert ws['A2'].value is not None  # project_name
        assert ws['B2'].value is not None  # project_slug
        assert ws['C2'].value is not None  # description

    # ========================================================================
    # AUTOMATED PROJECT CREATION ENDPOINT TESTS
    # ========================================================================

    def test_create_project_from_spreadsheet_success(self, test_app, auth_headers, clean_database):
        """Test successful automated project creation"""
        project_data = {
            'project_name': 'Python Web Development',
            'project_slug': 'python-web-dev',
            'description': 'Comprehensive training',
            'tracks': ['Backend', 'Frontend', 'Database'],
            'students': [
                {'email': 'student1@example.com', 'name': 'Alice Johnson'},
                {'email': 'student2@example.com', 'name': 'Bob Smith'}
            ],
            'instructors': [
                {'email': 'instructor@example.com', 'name': 'Dr. Sarah Chen'}
            ]
        }

        response = test_app.post(
            "/api/v1/projects/create-from-spreadsheet",
            headers=auth_headers,
            json=project_data
        )

        assert response.status_code == 200
        result = response.json()
        assert result['success'] is True
        assert result['project_name'] == 'Python Web Development'
        assert result['project_id'] is not None
        assert result['tracks_created'] == 3
        assert 'processing_time_ms' in result

    def test_create_project_minimal_data(self, test_app, auth_headers, clean_database):
        """Test project creation with minimal data"""
        project_data = {
            'project_name': 'Minimal Project',
            'project_slug': 'minimal-project',
            'description': 'Test description'
        }

        response = test_app.post(
            "/api/v1/projects/create-from-spreadsheet",
            headers=auth_headers,
            json=project_data
        )

        assert response.status_code == 200
        result = response.json()
        assert result['success'] is True
        assert result['tracks_created'] == 0
        assert result['students_enrolled'] == 0
        assert result['instructors_assigned'] == 0

    def test_create_project_with_students_no_names(self, test_app, auth_headers, clean_database):
        """Test project creation with students without names"""
        project_data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description',
            'tracks': ['Backend'],
            'students': [
                {'email': 'student1@example.com'},
                {'email': 'student2@example.com'}
            ]
        }

        response = test_app.post(
            "/api/v1/projects/create-from-spreadsheet",
            headers=auth_headers,
            json=project_data
        )

        assert response.status_code == 200
        result = response.json()
        assert result['success'] is True

    def test_create_project_without_auth(self, test_app, auth_headers, clean_database):
        """Test project creation without authentication"""
        project_data = {
            'project_name': 'Python Training',
            'project_slug': 'python-training',
            'description': 'Test description'
        }

        response = test_app.post(
            "/api/v1/projects/create-from-spreadsheet",
            json=project_data
        )

        # Should require authentication
        assert response.status_code in [401, 403, 422]

    # ========================================================================
    # END-TO-END WORKFLOW TESTS
    # ========================================================================

    def test_complete_workflow_import_then_create(self, test_app, auth_headers, clean_database):
        """Test complete workflow: import spreadsheet then create project"""
        # Step 1: Import spreadsheet
        data = {
            'project_name': 'Complete Workflow Test',
            'project_slug': 'complete-workflow-test',
            'description': 'Testing complete workflow',
            'tracks': 'Track 1, Track 2',
            'student_emails': 'student@example.com',
            'student_names': 'Test Student'
        }

        xlsx_bytes = self.create_test_xlsx(data)

        import_response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            headers=auth_headers,
            files={"file": ("project.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert import_response.status_code == 200
        parsed_data = import_response.json()

        # Step 2: Create project from parsed data
        create_response = test_app.post(
            "/api/v1/projects/create-from-spreadsheet",
            headers=auth_headers,
            json=parsed_data
        )

        assert create_response.status_code == 200
        result = create_response.json()
        assert result['success'] is True
        assert result['project_name'] == 'Complete Workflow Test'
        assert result['tracks_created'] == 2

    def test_download_template_and_import(self, test_app, auth_headers, clean_database):
        """Test download template then import it back"""
        # Step 1: Download template
        template_response = test_app.get(
            "/api/v1/projects/template",
            headers=auth_headers
        )

        assert template_response.status_code == 200
        template_bytes = template_response.content

        # Step 2: Import the template (which has example data)
        import_response = test_app.post(
            "/api/v1/projects/import-spreadsheet",
            headers=auth_headers,
            files={"file": ("project.xlsx", template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert import_response.status_code == 200
        result = import_response.json()
        assert result['success'] is True
        assert result['project_name'] == 'Python Web Development Training'


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
