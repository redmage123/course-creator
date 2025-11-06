"""
Project Spreadsheet Parser Service

This service provides functionality to parse project data from various
spreadsheet formats (XLSX, XLS, ODS) for automated project creation.

BUSINESS CONTEXT:
Organization admins can upload spreadsheets containing project information
to quickly create projects with tracks. This service extracts and normalizes
project data from different file formats.

SUPPORTED FORMATS:
- XLSX (Microsoft Excel - Excel 2007+)
- XLS (Microsoft Excel - Legacy)
- ODS (LibreOffice Calc)

REQUIRED COLUMNS:
- project_name (REQUIRED): Project display name
- project_slug (REQUIRED): URL-friendly identifier (lowercase, hyphens)
- description (REQUIRED): Project description

OPTIONAL COLUMNS:
- start_date (OPTIONAL): Project start date (YYYY-MM-DD format)
- end_date (OPTIONAL): Project end date (YYYY-MM-DD format)
- tracks (OPTIONAL): Comma-separated list of track names
- student_emails (OPTIONAL): Comma-separated student email addresses
- student_names (OPTIONAL): Comma-separated student names (parallel to emails)
- instructor_emails (OPTIONAL): Comma-separated instructor email addresses
- instructor_names (OPTIONAL): Comma-separated instructor names (parallel to emails)

USAGE EXAMPLE:
    parser = ProjectSpreadsheetParser()
    project_data = parser.parse_file(file_bytes, 'project.xlsx')
    print(project_data['project_name'])
"""
import pandas as pd
import io
from typing import Dict, List
from datetime import datetime
import logging
import openpyxl
from openpyxl import Workbook

logger = logging.getLogger(__name__)


class ProjectSpreadsheetParser:
    """
    Service for parsing project data from spreadsheet files.

    BUSINESS REQUIREMENTS:
    - Support multiple Excel formats (XLSX, XLS, ODS)
    - Extract project data with flexible column mapping
    - Handle missing optional columns gracefully
    - Validate required columns are present
    - Provide descriptive error messages
    - Generate template spreadsheets for download
    """

    REQUIRED_COLUMNS = ['project_name', 'project_slug', 'description']
    OPTIONAL_COLUMNS = ['start_date', 'end_date', 'tracks', 'student_emails', 'student_names', 'instructor_emails', 'instructor_names']

    def parse_xlsx(self, xlsx_bytes: bytes) -> Dict:
        """
        Parse project data from XLSX/XLS binary content.

        BUSINESS LOGIC:
        - Read XLSX bytes into pandas DataFrame using openpyxl engine
        - Validate required columns exist
        - Extract project data from first row
        - Parse dates and track lists
        - Apply default values for missing optional fields

        Args:
            xlsx_bytes: XLSX file content as bytes

        Returns:
            Dictionary containing project data

        Raises:
            ValueError: If XLSX is empty or missing required columns
        """
        if not xlsx_bytes:
            raise ValueError("Empty spreadsheet")

        try:
            # Parse XLSX content
            df = pd.read_excel(io.BytesIO(xlsx_bytes), engine='openpyxl')

            # Validate and process
            return self._process_dataframe(df)

        except Exception as e:
            logger.error(f"Error parsing XLSX: {e}")
            raise ValueError(f"Failed to parse XLSX: {str(e)}")

    def parse_ods(self, ods_bytes: bytes) -> Dict:
        """
        Parse project data from ODS binary content.

        BUSINESS LOGIC:
        - Read ODS bytes into pandas DataFrame using odf engine
        - Validate required columns exist
        - Extract project data from first row
        - Parse dates and track lists
        - Apply default values for missing optional fields

        Args:
            ods_bytes: ODS file content as bytes

        Returns:
            Dictionary containing project data

        Raises:
            ValueError: If ODS is empty or missing required columns
        """
        if not ods_bytes:
            raise ValueError("Empty spreadsheet")

        try:
            # Parse ODS content
            df = pd.read_excel(io.BytesIO(ods_bytes), engine='odf')

            # Validate and process
            return self._process_dataframe(df)

        except Exception as e:
            logger.error(f"Error parsing ODS: {e}")
            raise ValueError(f"Failed to parse ODS: {str(e)}")

    def _process_dataframe(self, df: pd.DataFrame) -> Dict:
        """
        Process pandas DataFrame into project dictionary.

        BUSINESS LOGIC:
        - Validate required columns are present
        - Extract data from first row (single project per file)
        - Parse dates into ISO format
        - Split comma-separated tracks into list
        - Clean and normalize data

        Args:
            df: Pandas DataFrame containing project data

        Returns:
            Dictionary containing normalized project data

        Raises:
            ValueError: If required columns are missing
        """
        if df.empty:
            raise ValueError("Empty spreadsheet")

        # Validate required columns
        missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        # Extract first row (single project per file)
        row = df.iloc[0]

        # Build project data dictionary
        project_data = {
            'project_name': str(row['project_name']).strip() if pd.notna(row['project_name']) else None,
            'project_slug': str(row['project_slug']).strip().lower() if pd.notna(row['project_slug']) else None,
            'description': str(row['description']).strip() if pd.notna(row['description']) else None
        }

        # Validate required fields are not empty
        if not project_data['project_name']:
            raise ValueError("project_name cannot be empty")
        if not project_data['project_slug']:
            raise ValueError("project_slug cannot be empty")
        if not project_data['description']:
            raise ValueError("description cannot be empty")

        # Parse optional date fields
        if 'start_date' in df.columns and pd.notna(row['start_date']):
            project_data['start_date'] = self._parse_date(row['start_date'])

        if 'end_date' in df.columns and pd.notna(row['end_date']):
            project_data['end_date'] = self._parse_date(row['end_date'])

        # Parse optional tracks field (comma-separated list)
        if 'tracks' in df.columns and pd.notna(row['tracks']):
            tracks_str = str(row['tracks']).strip()
            if tracks_str:
                project_data['tracks'] = [t.strip() for t in tracks_str.split(',') if t.strip()]
            else:
                project_data['tracks'] = []
        else:
            project_data['tracks'] = []

        # Parse optional student fields (comma-separated lists)
        project_data['students'] = []
        if 'student_emails' in df.columns and pd.notna(row['student_emails']):
            emails_str = str(row['student_emails']).strip()
            if emails_str:
                student_emails = [e.strip() for e in emails_str.split(',') if e.strip()]

                # Parse student names if provided (parallel to emails)
                student_names = []
                if 'student_names' in df.columns and pd.notna(row['student_names']):
                    names_str = str(row['student_names']).strip()
                    if names_str:
                        student_names = [n.strip() for n in names_str.split(',') if n.strip()]

                # Build student list with email and optional name
                for idx, email in enumerate(student_emails):
                    student = {'email': email}
                    if idx < len(student_names):
                        student['name'] = student_names[idx]
                    project_data['students'].append(student)

        # Parse optional instructor fields (comma-separated lists)
        project_data['instructors'] = []
        if 'instructor_emails' in df.columns and pd.notna(row['instructor_emails']):
            emails_str = str(row['instructor_emails']).strip()
            if emails_str:
                instructor_emails = [e.strip() for e in emails_str.split(',') if e.strip()]

                # Parse instructor names if provided (parallel to emails)
                instructor_names = []
                if 'instructor_names' in df.columns and pd.notna(row['instructor_names']):
                    names_str = str(row['instructor_names']).strip()
                    if names_str:
                        instructor_names = [n.strip() for n in names_str.split(',') if n.strip()]

                # Build instructor list with email and optional name
                for idx, email in enumerate(instructor_emails):
                    instructor = {'email': email}
                    if idx < len(instructor_names):
                        instructor['name'] = instructor_names[idx]
                    project_data['instructors'].append(instructor)

        logger.info(f"Parsed project: {project_data['project_name']} with {len(project_data.get('tracks', []))} tracks, {len(project_data.get('students', []))} students, {len(project_data.get('instructors', []))} instructors")
        return project_data

    def _parse_date(self, date_value) -> str:
        """
        Parse date value into ISO format (YYYY-MM-DD).

        BUSINESS LOGIC:
        - Handle various date formats
        - Convert to ISO format string
        - Return None if parsing fails

        Args:
            date_value: Date value from spreadsheet (various types possible)

        Returns:
            ISO format date string (YYYY-MM-DD) or None
        """
        try:
            if isinstance(date_value, datetime):
                return date_value.strftime('%Y-%m-%d')
            elif isinstance(date_value, str):
                # Try parsing string date
                parsed_date = pd.to_datetime(date_value)
                return parsed_date.strftime('%Y-%m-%d')
            else:
                return None
        except Exception as e:
            logger.warning(f"Failed to parse date: {date_value}, error: {e}")
            return None

    def detect_format(self, filename: str) -> str:
        """
        Detect spreadsheet format from filename.

        BUSINESS LOGIC:
        - Use file extension to determine format

        Args:
            filename: Original filename with extension

        Returns:
            Format string: 'xlsx' or 'ods'

        Raises:
            ValueError: If format cannot be determined
        """
        extension = filename.lower().split('.')[-1]

        if extension in ['xlsx', 'xls']:
            return 'xlsx'
        elif extension == 'ods':
            return 'ods'
        else:
            raise ValueError(f"Unsupported file type: {extension}. Please use .xlsx, .xls, or .ods")

    def parse_file(self, file_content: bytes, filename: str) -> Dict:
        """
        Parse spreadsheet file automatically detecting format.

        BUSINESS LOGIC:
        - Detect file format from filename
        - Route to appropriate parser method
        - Return normalized project data

        Args:
            file_content: File content as bytes
            filename: Original filename with extension

        Returns:
            Dictionary containing project data

        Raises:
            ValueError: If format is unsupported or parsing fails
        """
        format_type = self.detect_format(filename)

        if format_type == 'xlsx':
            return self.parse_xlsx(file_content)
        elif format_type == 'ods':
            return self.parse_ods(file_content)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate a template XLSX file for project import.

        BUSINESS LOGIC:
        - Create workbook with proper headers
        - Add example row with sample data
        - Include column descriptions
        - Format for user-friendly editing

        Returns:
            XLSX file content as bytes

        TECHNICAL IMPLEMENTATION:
        - Uses openpyxl to create Excel file
        - Includes headers with required/optional indicators
        - Adds sample data row for guidance
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Import Template"

        # Define headers with descriptions
        headers = [
            'project_name',
            'project_slug',
            'description',
            'start_date',
            'end_date',
            'tracks',
            'student_emails',
            'student_names',
            'instructor_emails',
            'instructor_names'
        ]

        descriptions = [
            'Project Name (REQUIRED)',
            'URL Slug (REQUIRED, lowercase-hyphens)',
            'Project Description (REQUIRED)',
            'Start Date (OPTIONAL, YYYY-MM-DD)',
            'End Date (OPTIONAL, YYYY-MM-DD)',
            'Comma-separated Track Names (OPTIONAL)',
            'Comma-separated Student Emails (OPTIONAL)',
            'Comma-separated Student Names (OPTIONAL, parallel to emails)',
            'Comma-separated Instructor Emails (OPTIONAL)',
            'Comma-separated Instructor Names (OPTIONAL, parallel to emails)'
        ]

        # Write headers
        for col_idx, (header, description) in enumerate(zip(headers, descriptions), start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.comment = openpyxl.comments.Comment(description, "System")

        # Write example row
        example_data = [
            'Python Web Development Training',
            'python-web-dev-2024',
            'Comprehensive training program for building modern web applications with Python',
            '2024-01-15',
            '2024-06-30',
            'Backend Fundamentals, Frontend Basics, Database Design',
            'student1@example.com, student2@example.com',
            'Alice Johnson, Bob Smith',
            'instructor1@example.com',
            'Dr. Sarah Chen'
        ]

        for col_idx, value in enumerate(example_data, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("Generated project import template")
        return output.read()
